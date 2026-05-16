"""Build AICS Lead Tracker Excel file with dropdowns, conditional formatting, and summary sheet."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "C:/Users/ASUS VIVOBOOK 15/aics-web/partnership/02_lead_tracking.xlsx"

wb = Workbook()

# ============ LEADS SHEET ============
ws = wb.active
ws.title = "Leads"

HEADERS = [
    "Tanggal Outreach", "Nama Prospect", "Channel", "Link Profile",
    "Perusahaan", "Industri", "Posisi", "Pesan Dikirim (Singkat)",
    "Status", "Tanggal Reply", "Next Action", "Hari Trial",
    "Tanggal Bayar", "Plan Dipilih", "Komisi (Rp)", "Notes",
]
COL_WIDTHS = [16, 22, 14, 38, 24, 14, 18, 38, 16, 16, 28, 11, 14, 18, 14, 38]

# Header styling
header_font = Font(name="Arial", bold=True, color="FAFAF7", size=11)
header_fill = PatternFill("solid", start_color="1C1917")
header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="E7E5E4")
header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = header_border
    ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]

ws.row_dimensions[1].height = 32

# Sample data (10 rows)
SAMPLE = [
    ("2026-05-15", "Budi Santoso", "LinkedIn", "https://linkedin.com/in/budi-santoso",
     "Toko Mawar Online", "E-commerce", "Owner", "Cold DM offer free audit",
     "Sent", None, "Follow up 2026-05-17", None, None, "-", 0, "First cold DM"),
    ("2026-05-15", "Siti Rahayu", "WhatsApp", "wa.me/628123456789",
     "PT Maju Bersama", "SaaS", "Founder", "Warm intro from Andi",
     "Replied", "2026-05-15", "Schedule Zoom call", None, None, "-", 0, "Interested in trial"),
    ("2026-05-16", "Pak Hartono", "Instagram DM", "instagram.com/hartono.id",
     "Hartono Coffee", "F&B", "Owner", "Personalized about review pain",
     "Opened", None, "Wait 3 days", None, None, "-", 0, ""),
    ("2026-05-16", "Dewi Anggraini", "LinkedIn", "https://linkedin.com/in/dewi-anggraini",
     "Sociolla", "Beauty", "Head of CX", "Comment di post mereka dulu",
     "In Discussion", "2026-05-16", "Send proposal", None, None, "-", 0, "Big brand interest"),
    ("2026-05-17", "Rizki Pratama", "LinkedIn", "https://linkedin.com/in/rizki-pratama",
     "Janji Jiwa Coffee", "F&B", "Marketing", "Refer audit gratis",
     "Trial Signup", "2026-05-17", "Day 7 check-in", 7, None, "-", 0, "Onboarded"),
    ("2026-05-18", "Andi Wijaya", "Email", "andi@kopikenangan.com",
     "Kopi Kenangan", "F&B", "Operations", "Cold email about CS",
     "Lost", "2026-05-19", "Closed - they use Cekat", None, None, "-", 0, "Already using Cekat"),
    ("2026-05-18", "Maya Putri", "LinkedIn", "https://linkedin.com/in/maya-putri",
     "Hangry", "F&B", "Co-founder", "Mutual connection",
     "Replied", "2026-05-18", "Send demo video", None, None, "-", 0, ""),
    ("2026-05-19", "Pak Bambang", "WhatsApp", "wa.me/628987654321",
     "UD Sejahtera", "Retail", "Owner", "Warm intro from cousin",
     "Demo Booked", "2026-05-19", "Zoom 2026-05-22 14:00", None, None, "-", 0, "Family referral"),
    ("2026-05-20", "Ningrum", "LinkedIn", "https://linkedin.com/in/ningrum",
     "Eatlah", "F&B", "Operations", "Reply on her post",
     "In Discussion", "2026-05-20", "Follow up tomorrow", None, None, "-", 0, ""),
    ("2026-05-22", "Pak Hartono", "WhatsApp", "wa.me/628333222111",
     "Hartono Coffee", "F&B", "Owner", "Follow up dari IG",
     "Paid", "2026-05-25", "Activate plan", 14, "2026-05-25", "Pro (Rp 299rb)", 30000,
     "First paying! Pro Rp 299rb x 10% = Rp 30k"),
]

cell_font = Font(name="Arial", size=10.5)
data_align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

for row_idx, row_data in enumerate(SAMPLE, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.alignment = data_align_top
        cell.border = header_border
    ws.row_dimensions[row_idx].height = 38

# Freeze top row
ws.freeze_panes = "A2"

# ============ DATA VALIDATION (Dropdowns) ============
# Apply dropdowns to rows 2-1000

dv_channel = DataValidation(
    type="list",
    formula1='"LinkedIn,WhatsApp,Email,Instagram DM,Komunitas,Facebook Group,TikTok DM,Referral"',
    allow_blank=True
)
dv_channel.add("C2:C1000")
ws.add_data_validation(dv_channel)

dv_industri = DataValidation(
    type="list",
    formula1='"E-commerce,SaaS,F&B,Edu,Beauty,Retail,Fashion,Health,Travel,Other"',
    allow_blank=True
)
dv_industri.add("F2:F1000")
ws.add_data_validation(dv_industri)

dv_status = DataValidation(
    type="list",
    formula1='"Sent,Opened,Replied,In Discussion,Demo Booked,Demo Done,Trial Signup,Paid,Lost"',
    allow_blank=True
)
dv_status.add("I2:I1000")
ws.add_data_validation(dv_status)

dv_plan = DataValidation(
    type="list",
    formula1='"-,Lite (Rp 99rb),Pro (Rp 299rb),Business (Rp 999rb),Custom"',
    allow_blank=True
)
dv_plan.add("N2:N1000")
ws.add_data_validation(dv_plan)

# ============ CONDITIONAL FORMATTING on Status column ============
status_styles = [
    ("Paid", "DCFCE7", "15803D"),
    ("Lost", "FEE2E2", "DC2626"),
    ("Demo Booked", "FEF3C7", "D97706"),
    ("Trial Signup", "FEF3C7", "D97706"),
    ("Demo Done", "DBEAFE", "2563EB"),
    ("In Discussion", "DBEAFE", "2563EB"),
    ("Replied", "EDE9FE", "7C3AED"),
    ("Sent", "F5F4F1", "44403C"),
    ("Opened", "F5F4F1", "44403C"),
]

for status, bg, fg in status_styles:
    rule = CellIsRule(
        operator="equal",
        formula=[f'"{status}"'],
        fill=PatternFill("solid", start_color=bg),
        font=Font(name="Arial", size=10.5, color=fg, bold=True),
    )
    ws.conditional_formatting.add("I2:I1000", rule)

# Format Komisi column as currency
for r in range(2, 1001):
    cell = ws.cell(row=r, column=15)  # Column O (Komisi)
    cell.number_format = '"Rp "#,##0;[Red]"-Rp "#,##0;"-"'

# ============ SUMMARY SHEET ============
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 22

title = ws2.cell(row=1, column=1, value="AICS Lead Tracking Summary")
title.font = Font(name="Arial", bold=True, size=16, color="1C1917")
ws2.row_dimensions[1].height = 28

ws2.cell(row=2, column=1, value="Auto-updated dari sheet 'Leads'").font = Font(name="Arial", italic=True, color="78716C", size=10)

# Main KPIs
KPIS = [
    ("Total Outreach", "=COUNTA(Leads!A2:A1000)"),
    ("Total Replied or Beyond", '=COUNTIF(Leads!I2:I1000,"Replied")+COUNTIF(Leads!I2:I1000,"In Discussion")+COUNTIF(Leads!I2:I1000,"Demo Booked")+COUNTIF(Leads!I2:I1000,"Demo Done")+COUNTIF(Leads!I2:I1000,"Trial Signup")+COUNTIF(Leads!I2:I1000,"Paid")+COUNTIF(Leads!I2:I1000,"Lost")'),
    ("Reply Rate", "=IFERROR(B5/B4, 0)"),
    ("Demo Booked / Done", '=COUNTIF(Leads!I2:I1000,"Demo Booked")+COUNTIF(Leads!I2:I1000,"Demo Done")'),
    ("Trial Signup or Beyond", '=COUNTIF(Leads!I2:I1000,"Trial Signup")+COUNTIF(Leads!I2:I1000,"Paid")'),
    ("Paid Customer", '=COUNTIF(Leads!I2:I1000,"Paid")'),
    ("Total Komisi (Rp)", "=SUM(Leads!O2:O1000)"),
]
section_font = Font(name="Arial", bold=True, size=11, color="C2410C")
kpi_label_font = Font(name="Arial", size=11, color="44403C")
kpi_value_font = Font(name="Arial", bold=True, size=12, color="1C1917")

for i, (label, formula) in enumerate(KPIS):
    r = 4 + i
    label_cell = ws2.cell(row=r, column=1, value=label)
    label_cell.font = kpi_label_font
    value_cell = ws2.cell(row=r, column=2, value=formula)
    value_cell.font = kpi_value_font
    if "Rate" in label:
        value_cell.number_format = "0.0%"
    elif "Komisi" in label:
        value_cell.number_format = '"Rp "#,##0'
    else:
        value_cell.number_format = "0"

# By Channel breakdown
ws2.cell(row=13, column=1, value="📊 By Channel").font = section_font
CHANNELS = ["LinkedIn", "WhatsApp", "Email", "Instagram DM", "Komunitas", "Facebook Group", "TikTok DM", "Referral"]
for i, ch in enumerate(CHANNELS):
    r = 14 + i
    ws2.cell(row=r, column=1, value=ch).font = kpi_label_font
    cell = ws2.cell(row=r, column=2, value=f'=COUNTIF(Leads!C2:C1000,"{ch}")')
    cell.font = kpi_value_font
    cell.number_format = "0"

# By Industri breakdown
ws2.cell(row=23, column=1, value="🏢 By Industri").font = section_font
INDUSTRI = ["E-commerce", "SaaS", "F&B", "Edu", "Beauty", "Retail", "Fashion", "Health", "Travel", "Other"]
for i, ind in enumerate(INDUSTRI):
    r = 24 + i
    ws2.cell(row=r, column=1, value=ind).font = kpi_label_font
    cell = ws2.cell(row=r, column=2, value=f'=COUNTIF(Leads!F2:F1000,"{ind}")')
    cell.font = kpi_value_font
    cell.number_format = "0"

# By Plan breakdown
ws2.cell(row=35, column=1, value="💰 By Plan").font = section_font
PLANS = [
    ("Lite (Rp 99rb)", "Lite (Rp 99rb)", 99000),
    ("Pro (Rp 299rb)", "Pro (Rp 299rb)", 299000),
    ("Business (Rp 999rb)", "Business (Rp 999rb)", 999000),
    ("Custom", "Custom", 15000000),
]
for i, (label, exact_value, price) in enumerate(PLANS):
    r = 36 + i
    ws2.cell(row=r, column=1, value=label).font = kpi_label_font
    cell = ws2.cell(row=r, column=2, value=f'=COUNTIF(Leads!N2:N1000,"{exact_value}")')
    cell.font = kpi_value_font
    cell.number_format = "0"

# MRR estimate
ws2.cell(row=41, column=1, value="Estimated MRR").font = section_font
mrr_cell = ws2.cell(row=42, column=1, value="Monthly Recurring Revenue (Rp)")
mrr_cell.font = kpi_label_font
mrr_value = ws2.cell(row=42, column=2,
    value='=COUNTIF(Leads!N2:N1000,"Lite (Rp 99rb)")*99000+COUNTIF(Leads!N2:N1000,"Pro (Rp 299rb)")*299000+COUNTIF(Leads!N2:N1000,"Business (Rp 999rb)")*999000'
)
mrr_value.font = Font(name="Arial", bold=True, size=14, color="C2410C")
mrr_value.number_format = '"Rp "#,##0'

# Borders for summary
for r in list(range(4, 11)) + list(range(13, 22)) + list(range(23, 34)) + list(range(35, 40)) + [41, 42]:
    for c in range(1, 3):
        ws2.cell(row=r, column=c).border = header_border

# ============ SAVE ============
wb.save(OUT)
print(f"Saved: {OUT}")
